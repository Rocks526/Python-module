from openpyxl import Workbook
import datetime

# 实例化
wb = Workbook()

# 激活 worksheet
ws = wb.active

# 创建表
# 方式一：插入到最后(default)
ws1 = wb.create_sheet("Sheet1")
# 方式二：插入到第二个
ws2 = wb.create_sheet("Sheet2", 1)

# 插入数据
# 方式一：数据可以直接分配到单元格中(可以输入公式)
ws['A1'] = 42
# 方式二：可以附加行，从第一列开始附加(从最下方空白处，最左开始)(可以输入多行)
ws.append([1, 2, 3])
# 方式三：Python 类型会被自动转换
ws['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")

# 访问单元格
# 方法一
c = ws['A4']
# 方法二：row 行；column 列
d = ws.cell(row=4, column=2, value=10)
# 方法三：只要访问就创建
for i in range(1, 101):
    for j in range(1, 101):
        ws.cell(row=i, column=j)

# 选择表
ws3 = wb["Sheet2"]
ws4 = wb.get_sheet_by_name("Sheet1")

# 查看表名
print(wb.sheetnames)

# 获得最大列和最大行
print(ws.max_row)
print(ws.max_column)

# 保存数据
wb.save('demo.xlsx')